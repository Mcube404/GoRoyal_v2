import openpyxl
import json
import re

def normalize_key(name):
    return re.sub(r'[^a-zA-Z0-9]+', '-', name.lower()).strip('-')

def parse_excel(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    
    # Skip 'Main categories' if we only need the details from other sheets
    # But let's see what's in 'Main categories' just in case
    main_cats = []
    if 'Main categories' in wb.sheetnames:
        sheet = wb['Main categories']
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if row[0]:
                main_cats.append(row[0])
    
    result = {}
    
    # Mapping sheet names to JSON keys
    name_map = {
        'Automobiles': 'automobiles',
        'watches': 'watches',
        'Electronics': 'electronics',
        'Home Appliances': 'home-appliances',
        'Bags and footwear': 'bags-footwear',
        'Home decor': 'home-decor',
        'fitness': 'fitness',
        'Sports': 'sports',
        'kids and toddlers': 'kids-toddlers',
        'pet supplies': 'pet-supplies',
        'mens': 'men',
        'womens': 'women',
        'Musical Instruments': 'musical-instruments'
    }
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Main categories':
            continue
            
        json_key = name_map.get(sheet_name)
        if not json_key:
            json_key = normalize_key(sheet_name)
            
        sheet = wb[sheet_name]
        data = list(sheet.iter_rows(values_only=True))
        if not data:
            continue
            
        headers = data[0]
        # Skip if first row is empty
        if not any(headers):
            continue
            
        subcategories = []
        for i, h in enumerate(headers):
            if h is None:
                continue
            
            items = []
            for r in range(1, len(data)):
                val = data[r][i]
                if val is not None:
                    items.append(str(val).strip())
            
            subcategories.append({
                "name": h,
                "img": f"Images/{normalize_key(h)}.png", # Placeholder/derived image path
                "topRatedUrl": "#",
                "sponsoredUrl": "#",
                "items": items
            })
        
        result[json_key] = subcategories
        
    return result

if __name__ == "__main__":
    data = parse_excel('goroyal_categories_v1 (2).xlsx')
    with open('parsed_categories.json', 'w') as f:
        json.dump(data, f, indent=4)
    print("Done")
